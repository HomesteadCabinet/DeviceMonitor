import os
import platform
import subprocess
import smtplib
import local_config
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
import requests
import time
import socket

# Constants for Online and Offline statuses
ONLINE = "Online"
OFFLINE = "Offline"

# List of devices to monitor with named URLs, IPs, directories, and ports
devices = local_config.devices

# Email settings
email_header = "Network Monitoring System"
sender_email = local_config.sender_email
sender_name = local_config.sender_name
receiver_emails = local_config.receiver_emails
email_password = local_config.email_password
smtp_server = local_config.smtp_server
smtp_port = local_config.smtp_port

# Threshold for slow response time (default 5 seconds = 5000 milliseconds)
RESPONSE_TIME_THRESHOLD = 5000

# Path to Excel log file
TEMP_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(TEMP_DIR, "device_status_log.xlsx")


def initialize_log():
    """Initialize the Excel log file if it doesn't exist."""
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Device Status"

        # Updated Header row with "Value" moved to column D and "Previous Status" after "Status"
        headers = [
            "Device Name      ",
            "Resource         ",
            "Type",
            "Value                   ",
            "Status",
            "Previous Status",
            "Last Checked       ",
            "Offline Since      ",
            "Online Since       "
        ]
        ws.append(headers)

        # Resize columns to fit the data
        bold_font = Font(bold=True, name="Arial", size=11)
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"].font = bold_font
            ws.column_dimensions[col_letter].width = max(len(header) + 8, 15)  # Adding some padding

        wb.save(LOG_FILE)


def update_device_status(device_name, resource_name, resource_type, status, value):
    """Update or insert device status in the Excel log file."""
    wb = load_workbook(LOG_FILE)
    ws = wb.active

    current_time = datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    updated = False
    arial_font = Font(name='Arial', size=10)
    red_font = Font(name='Arial', size=10, color="CC2345")

    # Iterate over rows to find the matching device/resource and update the status
    for row in ws.iter_rows(min_row=2, values_only=False):
        if (row[0].value == device_name and row[1].value == resource_name and row[2].value == resource_type):
            # Update existing entry
            previous_status = row[4].value  # Status is now column E (index 4)
            row[5].value = previous_status  # Previous Status is column F (index 5)
            row[3].value = value            # Value is column D (index 3)
            row[4].value = status            # Update to new status
            row[6].value = current_time      # Last Checked is column G (index 6)

            # Update Offline Since and Online Since
            if previous_status != status:
                if status == OFFLINE:
                    row[7].value = current_time    # Offline Since is column H (index 7)
                    row[8].value = ""              # Clear Online Since
                elif status == ONLINE:
                    row[8].value = current_time    # Online Since is column I (index 8)
                    row[7].value = ""              # Clear Offline Since
            updated = True
            break

    if not updated:
        # Add a new entry if device/resource not found
        offline_since = current_time if status == OFFLINE else ""
        online_since = current_time if status == ONLINE else ""
        ws.append([
            device_name,
            resource_name,
            resource_type,
            value,
            status,
            "",
            current_time,
            offline_since,
            online_since
        ])

    # Apply Arial font to all cells and red font to entire row if status is offline
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[4].value == OFFLINE:  # Column E is the status column
            for cell in row:
                cell.font = red_font
        else:
            for cell in row:
                cell.font = arial_font

    wb.save(LOG_FILE)


def get_previous_status(device_name, resource_name, resource_type):
    """Retrieve the previous status of a device/resource from the Excel log file."""
    wb = load_workbook(LOG_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if (row[0] == device_name and row[1] == resource_name and row[2] == resource_type):
            return row[4]  # Return the last known status from column E
    return None  # No previous status found


def ping_device(ip_info, device_name):
    """Ping a device and return its status and response time."""
    ip = ip_info['value']
    print(f"Starting ping check for {device_name} ({ip_info['name']}) - {ip}")
    try:
        start_time = time.time()
        command = ["ping", "-n", "1", ip] if platform.system().lower() == "windows" else ["ping", "-c", "1", ip]
        ping = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        end_time = (time.time() - start_time) * 1000  # Convert to milliseconds
        if ping.returncode == 0:
            print(f"    {device_name} ({ip_info['name']}) - {ONLINE} ({end_time:.2f}ms)")
            return ONLINE, end_time
        else:
            print(f"    {device_name} ({ip_info['name']}) - {OFFLINE}")
            return OFFLINE, None
    except Exception as e:
        print(f"    {device_name} ({ip_info['name']}) - {OFFLINE} - Error: {e}")
        return OFFLINE, None


def check_port(ip_info, device_name):
    """Check specified ports and return status."""
    ip = ip_info['value']
    if 'ports' not in ip_info:
        return ONLINE, None  # If no ports specified, assume online

    port_statuses = []
    for port in ip_info['ports']:
        print(f"Starting port check for {device_name} ({ip_info['name']}) - {ip}:{port}")
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)  # 3 seconds timeout
            start_time = time.time()
            result = sock.connect_ex((ip, port))
            end_time = (time.time() - start_time) * 1000  # Convert to milliseconds
            if result == 0:
                print(f"    {device_name} ({ip_info['name']}) Port {port} - {ONLINE} ({end_time:.2f}ms)")
                port_statuses.append((port, ONLINE, end_time))
            else:
                print(f"    {device_name} ({ip_info['name']}) Port {port} - {OFFLINE}")
                port_statuses.append((port, OFFLINE, None))
            sock.close()
        except Exception as e:
            print(f"    {device_name} ({ip_info['name']}) Port {port} - {OFFLINE} - Error: {e}")
            sock.close()
            port_statuses.append((port, OFFLINE, None))
    # For simplicity, return the first port status. Adjust if needed.
    if port_statuses:
        return port_statuses[0][1], port_statuses[0][2]
    return OFFLINE, None


def check_http(url_info, device_name):
    """Check HTTP response and return its status and response time."""
    url = url_info['value']
    print(f"Starting HTTP check for {device_name} ({url_info['name']}) - {url}")
    try:
        start_time = time.time()
        response = requests.get(url, timeout=5)
        end_time = (time.time() - start_time) * 1000  # Convert to milliseconds
        if response.status_code == 200:
            print(f"    {device_name} ({url_info['name']}) - {ONLINE} ({end_time:.2f}ms)")
            return ONLINE, end_time
        else:
            print(f"    {device_name} ({url_info['name']}) - {OFFLINE}")
            return OFFLINE, None
    except Exception as e:
        print(f"    {device_name} ({url_info['name']}) - {OFFLINE} - Error: {e}")
        return OFFLINE, None


def check_directory(directory_info, device_name):
    """Check if directory exists and return its status."""
    directory = directory_info['value']
    print(f"Starting directory check for {device_name} ({directory_info['name']}) - {directory}")
    try:
        if os.path.exists(directory):
            print(f"    {device_name} ({directory_info['name']}) - {ONLINE}")
            return ONLINE, None  # No response time for directories
        else:
            print(f"    {device_name} ({directory_info['name']}) - {OFFLINE}")
            return OFFLINE, None
    except Exception as e:
        print(f"    {device_name} ({directory_info['name']}) - {OFFLINE} - Error: {e}")
        return OFFLINE, None


def send_summary_email(offline_devices, online_devices):
    """Send a single email with a summary of offline and online devices, including response times."""
    if not offline_devices and not online_devices:
        print("No changes in status since last run... all done.")
        return

    offline_count = len(offline_devices)
    online_count = len(online_devices)
    body = ""

    subject_parts = []
    if offline_count > 0:
        offline_label = "Device" if offline_count == 1 else "Devices"
        subject_parts.append(f"{offline_count} New Offline {offline_label}")
    if online_count > 0:
        online_label = "Device" if online_count == 1 else "Devices"
        subject_parts.append(f"{online_count} New Online {online_label}")

    subject = "Devices"
    if subject_parts:
        subject += " - " + " and ".join(subject_parts)

    if offline_devices:
        body += "Devices that went offline:\n"
        for device, resource, value, response_time in offline_devices:
            if response_time:
                append = f" - {response_time:.2f}ms"
            else:
                append = ""
            body += f"{device} - {resource} ({value}){append}\n"

    if online_devices:
        body += "\nDevices that came back online:\n"
        for device, resource, value, response_time in online_devices:
            if response_time:
                append = f" - {response_time:.2f}ms"
            else:
                append = ""
            body += f"{device} - {resource} ({value}){append}\n"

    send_email(subject, body)


def send_email(subject, body):
    """Send an email to notify the recipient of status changes."""
    print(f"Sending email to {', '.join(receiver_emails)}")
    print(f"Subject: {subject}")
    print(f"Body:\n{body}")
    message = MIMEMultipart()
    message["From"] = formataddr((sender_name, sender_email))
    message["To"] = ", ".join(receiver_emails)
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain"))

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, email_password)
        server.sendmail(sender_email, receiver_emails, message.as_string())
        server.quit()
        print(f"Email sent to {', '.join(receiver_emails)}")
    except Exception as e:
        print(f"Failed to send email: {e}")


def check_devices():
    """Check the status of all devices and collect any that changed status."""
    offline_devices = []
    online_devices = []

    for device_name, resources in devices.items():
        # Check URLs
        if "urls" in resources:
            for url_info in resources["urls"]:
                current_status, response_time = check_http(url_info, device_name)
                previous_status = get_previous_status(device_name, url_info['name'], "URL")

                # Update device status before handling status changes
                update_device_status(device_name, url_info['name'], "URL", current_status, url_info['value'])

                # Handle the case when it's the first run (no previous status)
                if previous_status is None:
                    if current_status == OFFLINE:
                        offline_devices.append((device_name, url_info['name'], url_info['value'], response_time))
                    elif current_status == ONLINE:
                        online_devices.append((device_name, url_info['name'], url_info['value'], response_time))
                    continue

                # Handle status transitions
                if previous_status == ONLINE and current_status == OFFLINE:
                    offline_devices.append((device_name, url_info['name'], url_info['value'], response_time))
                elif previous_status == OFFLINE and current_status == ONLINE:
                    online_devices.append((device_name, url_info['name'], url_info['value'], response_time))

        # Check IPs
        if "ips" in resources:
            for ip_info in resources["ips"]:
                if ip_info.get('ports'):
                    current_status, response_time = check_port(ip_info, device_name)
                else:
                    current_status, response_time = ping_device(ip_info, device_name)
                previous_status = get_previous_status(device_name, ip_info['name'], "IP")

                # Update device status before handling status changes
                update_device_status(device_name, ip_info['name'], "IP", current_status, ip_info['value'])

                # Handle the case when it's the first run (no previous status)
                if previous_status is None:
                    if current_status == OFFLINE:
                        offline_devices.append((device_name, ip_info['name'], ip_info['value'], response_time))
                    elif current_status == ONLINE:
                        online_devices.append((device_name, ip_info['name'], ip_info['value'], response_time))
                    continue

                # Handle status transitions
                if previous_status == ONLINE and current_status == OFFLINE:
                    offline_devices.append((device_name, ip_info['name'], ip_info['value'], response_time))
                elif previous_status == OFFLINE and current_status == ONLINE:
                    online_devices.append((device_name, ip_info['name'], ip_info['value'], response_time))

        # Check directories
        if "directories" in resources:
            for directory_info in resources["directories"]:
                current_status, response_time = check_directory(directory_info, device_name)
                previous_status = get_previous_status(device_name, directory_info['name'], "Directory")

                # Update device status before handling status changes
                update_device_status(device_name, directory_info['name'], "Directory", current_status, directory_info['value'])

                # Handle the case when it's the first run (no previous status)
                if previous_status is None:
                    if current_status == OFFLINE:
                        offline_devices.append((device_name, directory_info['name'], directory_info['value'], response_time))
                    elif current_status == ONLINE:
                        online_devices.append((device_name, directory_info['name'], directory_info['value'], response_time))
                    continue

                # Handle status transitions
                if previous_status == ONLINE and current_status == OFFLINE:
                    offline_devices.append((device_name, directory_info['name'], directory_info['value'], response_time))
                elif previous_status == OFFLINE and current_status == ONLINE:
                    online_devices.append((device_name, directory_info['name'], directory_info['value'], response_time))

    return offline_devices, online_devices


initialize_log()
